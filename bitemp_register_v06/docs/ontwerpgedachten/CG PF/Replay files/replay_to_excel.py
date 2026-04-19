#!/usr/bin/env python3
"""
Converteer de CG Portfolio replay-files naar een Excel workbook.

Tabbladen:
  Referentielijsten:
    - Gemeente          (id, code, naam)
    - Domein            (id, naam, omschrijving)
    - ApiStandaard      (id, naam)
  Entiteiten:
    - Organisatie       (id, naam)
    - Persoon           (id, naam, email)
    - Initiatief        (brede tabel met alle GE-velden als kolommen)
  Koppelingen:
    - Contactpersoon    (organisatie → persoon link)

Buiten scope (per instructie): files 5b en 6.
"""

import json
import os
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent

FILES = [
    "1. Gemeenten CBS 2026.replay.json",
    "2. Domeinen vast 2026.replay.json",
    "3. API standaarden rationalisatie 2026.replay.json",
    "4. Intake Portfolio Common Ground 2.replay (zonder gemeenten).json",
    "5. PO email naar Persoon.Contactgegevens 2026.replay - zonder piet en test.json",
]

# ── helpers ──────────────────────────────────────────────────────────────────

def load_replay(filename):
    with open(SCRIPT_DIR / filename, encoding="utf-8") as f:
        return json.load(f)

def collect_opvoeren(data):
    """Yield (type_key, fields_dict) voor alle opvoer-wijzigingen."""
    for entry in data.get("entries", []):
        for w in entry.get("request_body", {}).get("wijzigingen", []):
            op = w.get("opvoer", {})
            for typ, fields in op.items():
                yield typ, fields

# ── data-extractie ───────────────────────────────────────────────────────────

def build_lookups():
    """Bouw lookup-dicts voor referentielijsten en entiteiten."""
    # Referentielijsten
    gemeenten = {}        # id → {naam, code}
    domeinen = {}         # id → {naam, omschrijving}
    apistandaarden = {}   # id → {naam}

    # Entiteiten
    organisaties = {}     # id → {naam}
    personen = {}         # id → {naam, email}

    # Initiatief child records, per initiatief_id
    initiatief_ids = set()
    planning = {}         # initiatief_id → dict
    product = {}          # initiatief_id → dict
    initiatief_aanvang = {}  # initiatief_id → datum
    initiatiefinfo = defaultdict(list)  # initiatief_id → [informatie]
    bijdragen = defaultdict(list)       # initiatief_id → [{type_bijdrage, schaal, toelichting}]
    initiatiefdomein = defaultdict(list)   # initiatief_id → [domein_id]
    initiatiefgemeente = defaultdict(list)  # initiatief_id → [{gemeente_id, rol}]
    initiatieforganisatie = defaultdict(list)  # initiatief_id → [{organisatie_id, rol}]
    initiatiefapistandaard = defaultdict(list)  # initiatief_id → [apistandaard_id]
    andereapistandaard = defaultdict(list)  # initiatief_id → [api_standaard]
    anderdomein = defaultdict(list)    # initiatief_id → [domein]
    andersdangemeente = defaultdict(list)  # initiatief_id → [tekst]
    betrokkenorganisatie = defaultdict(list)  # initiatief_id → [type]

    # Contactpersoon
    contactpersonen = []  # [{organisatie_id, persoon_id, rol}]

    for fname in FILES:
        data = load_replay(fname)
        for typ, fields in collect_opvoeren(data):
            if typ == "gemeente":
                gemeenten.setdefault(fields["id"], {})
            elif typ == "gemeentegegevens":
                gemeenten.setdefault(fields["gemeente_id"], {}).update(
                    {"naam": fields.get("naam", ""), "code": fields.get("code", "")}
                )
            elif typ == "domein":
                domeinen.setdefault(fields["id"], {})
            elif typ == "domeingegevens":
                domeinen.setdefault(fields["domein_id"], {}).update(
                    {"naam": fields.get("naam", ""), "omschrijving": fields.get("omschrijving", "")}
                )
            elif typ == "apistandaard":
                apistandaarden.setdefault(fields["id"], {})
            elif typ == "naam" and "apistandaard_id" in fields:
                apistandaarden.setdefault(fields["apistandaard_id"], {})["naam"] = fields.get("naam", "")

            elif typ == "organisatie":
                organisaties.setdefault(fields["id"], {})
            elif typ == "organisatienaam":
                organisaties.setdefault(fields["organisatie_id"], {})["naam"] = fields.get("naam", "")

            elif typ == "persoon":
                personen.setdefault(fields["id"], {})
            elif typ == "persoonnaam":
                personen.setdefault(fields["persoon_id"], {})["naam"] = fields.get("naam", "")
            elif typ == "persoonscontactgegevens" and "persoon_id" in fields:
                personen.setdefault(fields["persoon_id"], {})["email"] = fields.get("email", "")

            elif typ == "initiatief":
                initiatief_ids.add(fields["id"])
            elif typ == "initiatief_aanvang":
                initiatief_aanvang[fields["initiatief_id"]] = fields.get("datum", "")
            elif typ == "planning":
                planning[fields["initiatief_id"]] = {
                    k: v for k, v in fields.items() if k != "initiatief_id"
                }
            elif typ == "product":
                product[fields["initiatief_id"]] = {
                    k: v for k, v in fields.items() if k != "initiatief_id"
                }
            elif typ == "initiatiefinfo":
                initiatiefinfo[fields["initiatief_id"]].append(fields.get("informatie", ""))
            elif typ == "bijdrage":
                bijdragen[fields["initiatief_id"]].append({
                    k: v for k, v in fields.items() if k != "initiatief_id"
                })
            elif typ == "initiatiefdomein":
                initiatiefdomein[fields["initiatief_id"]].append(fields["domein_id"])
            elif typ == "initiatiefgemeente":
                initiatiefgemeente[fields["initiatief_id"]].append({
                    "gemeente_id": fields["gemeente_id"],
                    "rol": fields.get("rol", ""),
                })
            elif typ == "initiatieforganisatie":
                initiatieforganisatie[fields["initiatief_id"]].append({
                    "organisatie_id": fields["organisatie_id"],
                    "rol": fields.get("rol", ""),
                })
            elif typ == "initiatiefapistandaard":
                initiatiefapistandaard[fields["initiatief_id"]].append(fields["apistandaard_id"])
            elif typ == "andereapistandaard":
                andereapistandaard[fields["initiatief_id"]].append(fields.get("api_standaard", ""))
            elif typ == "anderdomein":
                anderdomein[fields["initiatief_id"]].append(fields.get("domein", ""))
            elif typ == "andersdangemeente":
                andersdangemeente[fields["initiatief_id"]].append(fields.get("andersDanGemeente", ""))
            elif typ == "betrokkenorganisatie":
                betrokkenorganisatie[fields["initiatief_id"]].append(fields.get("type", ""))
            elif typ == "contactpersoon":
                contactpersonen.append(dict(fields))

    return dict(
        gemeenten=gemeenten,
        domeinen=domeinen,
        apistandaarden=apistandaarden,
        organisaties=organisaties,
        personen=personen,
        initiatief_ids=sorted(initiatief_ids),
        planning=planning,
        product=product,
        initiatief_aanvang=initiatief_aanvang,
        initiatiefinfo=initiatiefinfo,
        bijdragen=bijdragen,
        initiatiefdomein=initiatiefdomein,
        initiatiefgemeente=initiatiefgemeente,
        initiatieforganisatie=initiatieforganisatie,
        initiatiefapistandaard=initiatiefapistandaard,
        andereapistandaard=andereapistandaard,
        anderdomein=anderdomein,
        andersdangemeente=andersdangemeente,
        betrokkenorganisatie=betrokkenorganisatie,
        contactpersonen=contactpersonen,
    )


# ── styling ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header(ws, ncols, fill=HEADER_FILL):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

def auto_width(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
            cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

def write_sheet(ws, headers, rows, fill=HEADER_FILL):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws, len(headers), fill)
    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ── Excel opbouw ─────────────────────────────────────────────────────────────

def build_excel(lookups):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Referentielijsten ────────────────────────────────────────────────

    # Gemeente
    ws = wb.create_sheet("Gemeente")
    rows = []
    for gid in sorted(lookups["gemeenten"]):
        g = lookups["gemeenten"][gid]
        rows.append([gid, g.get("code", ""), g.get("naam", "")])
    write_sheet(ws, ["id", "code", "naam"], rows, HEADER_FILL_GREEN)

    # Domein
    ws = wb.create_sheet("Domein")
    rows = []
    for did in sorted(lookups["domeinen"]):
        d = lookups["domeinen"][did]
        rows.append([did, d.get("naam", ""), d.get("omschrijving", "")])
    write_sheet(ws, ["id", "naam", "omschrijving"], rows, HEADER_FILL_GREEN)

    # ApiStandaard
    ws = wb.create_sheet("ApiStandaard")
    rows = []
    for aid in sorted(lookups["apistandaarden"]):
        a = lookups["apistandaarden"][aid]
        rows.append([aid, a.get("naam", "")])
    write_sheet(ws, ["id", "naam"], rows, HEADER_FILL_GREEN)

    # ── Entiteiten ───────────────────────────────────────────────────────

    # Organisatie
    ws = wb.create_sheet("Organisatie")
    rows = []
    for oid in sorted(lookups["organisaties"]):
        o = lookups["organisaties"][oid]
        rows.append([oid, o.get("naam", "")])
    write_sheet(ws, ["id", "naam"], rows, HEADER_FILL)

    # Persoon
    ws = wb.create_sheet("Persoon")
    rows = []
    for pid in sorted(lookups["personen"]):
        p = lookups["personen"][pid]
        rows.append([pid, p.get("naam", ""), p.get("email", "")])
    write_sheet(ws, ["id", "naam", "email"], rows, HEADER_FILL)

    # ── Initiatief (brede tabel) ─────────────────────────────────────────

    # Bepaal het max aantal bijdragen, domeinen, gemeenten, organisaties, apistandaarden
    max_bijdragen = max((len(lookups["bijdragen"][iid]) for iid in lookups["initiatief_ids"]), default=0)
    max_domeinen = max((len(lookups["initiatiefdomein"][iid]) for iid in lookups["initiatief_ids"]), default=0)
    max_gemeenten = max((len(lookups["initiatiefgemeente"][iid]) for iid in lookups["initiatief_ids"]), default=0)
    max_organisaties = max((len(lookups["initiatieforganisatie"][iid]) for iid in lookups["initiatief_ids"]), default=0)
    max_apistandaarden = max((len(lookups["initiatiefapistandaard"][iid]) for iid in lookups["initiatief_ids"]), default=0)
    max_andere_api = max((len(lookups["andereapistandaard"][iid]) for iid in lookups["initiatief_ids"]), default=0)
    max_ander_domein = max((len(lookups["anderdomein"][iid]) for iid in lookups["initiatief_ids"]), default=0)
    max_anders_gemeente = max((len(lookups["andersdangemeente"][iid]) for iid in lookups["initiatief_ids"]), default=0)
    max_betrokken = max((len(lookups["betrokkenorganisatie"][iid]) for iid in lookups["initiatief_ids"]), default=0)

    headers = [
        "initiatief.id",
        "initiatief_aanvang.datum",
        # product velden
        "product.naam",
        "product.omschrijving",
        "product.pitch",
        "product.type",
        "product.CG_laag",
        "product.website",
        "product.git_repo",
        # planning velden
        "planning.startdatum",
        "planning.ready_for_use",
        "planning.fase",
        "planning.planningsinfo",
        "planning.waar_tegenaan_gelopen",
        # initiatiefinfo
        "initiatiefinfo.informatie",
    ]

    # Bijdragen: type_bijdrage, schaal, toelichting per slot
    for i in range(1, max_bijdragen + 1):
        headers.append(f"bijdrage[{i}].type_bijdrage")
        headers.append(f"bijdrage[{i}].schaal")
        headers.append(f"bijdrage[{i}].toelichting")

    # Domeinen (resolved naam)
    for i in range(1, max_domeinen + 1):
        headers.append(f"initiatiefdomein[{i}].domein.naam")

    # Andere domeinen
    for i in range(1, max_ander_domein + 1):
        headers.append(f"anderdomein[{i}].domein")

    # Gemeenten (resolved naam)
    for i in range(1, max_gemeenten + 1):
        headers.append(f"initiatiefgemeente[{i}].gemeente.naam")
        headers.append(f"initiatiefgemeente[{i}].rol")

    # Anders dan gemeente
    for i in range(1, max_anders_gemeente + 1):
        headers.append(f"andersdangemeente[{i}]")

    # Organisaties (resolved naam + rol)
    for i in range(1, max_organisaties + 1):
        headers.append(f"initiatieforganisatie[{i}].organisatie.naam")
        headers.append(f"initiatieforganisatie[{i}].rol")

    # Betrokken organisatietypen
    for i in range(1, max_betrokken + 1):
        headers.append(f"betrokkenorganisatie[{i}].type")

    # API standaarden (resolved naam)
    for i in range(1, max_apistandaarden + 1):
        headers.append(f"initiatiefapistandaard[{i}].apistandaard.naam")

    # Andere API standaarden
    for i in range(1, max_andere_api + 1):
        headers.append(f"andereapistandaard[{i}].api_standaard")

    ws = wb.create_sheet("Initiatief")
    rows = []
    for iid in lookups["initiatief_ids"]:
        row = []
        row.append(iid)
        row.append(lookups["initiatief_aanvang"].get(iid, ""))

        # product
        prod = lookups["product"].get(iid, {})
        row.append(prod.get("naam", ""))
        row.append(prod.get("omschrijving", ""))
        row.append(prod.get("pitch", ""))
        row.append(prod.get("type", ""))
        row.append(prod.get("CG_laag", ""))
        row.append(prod.get("website", ""))
        row.append(prod.get("git_repo", ""))

        # planning
        plan = lookups["planning"].get(iid, {})
        row.append(plan.get("startdatum", ""))
        row.append(plan.get("ready_for_use", ""))
        row.append(plan.get("fase", ""))
        row.append(plan.get("planningsinfo", ""))
        row.append(plan.get("waar_tegenaan_gelopen", ""))

        # initiatiefinfo
        infos = lookups["initiatiefinfo"].get(iid, [])
        row.append(" | ".join(infos) if infos else "")

        # bijdragen
        bijs = lookups["bijdragen"].get(iid, [])
        for i in range(max_bijdragen):
            if i < len(bijs):
                row.append(bijs[i].get("type_bijdrage", ""))
                row.append(bijs[i].get("schaal", ""))
                row.append(bijs[i].get("toelichting", ""))
            else:
                row.extend(["", "", ""])

        # domeinen
        doms = lookups["initiatiefdomein"].get(iid, [])
        for i in range(max_domeinen):
            if i < len(doms):
                d = lookups["domeinen"].get(doms[i], {})
                row.append(d.get("naam", f"domein#{doms[i]}"))
            else:
                row.append("")

        # andere domeinen
        adoms = lookups["anderdomein"].get(iid, [])
        for i in range(max_ander_domein):
            row.append(adoms[i] if i < len(adoms) else "")

        # gemeenten
        gems = lookups["initiatiefgemeente"].get(iid, [])
        for i in range(max_gemeenten):
            if i < len(gems):
                g = lookups["gemeenten"].get(gems[i]["gemeente_id"], {})
                row.append(g.get("naam", f"gemeente#{gems[i]['gemeente_id']}"))
                row.append(gems[i].get("rol", ""))
            else:
                row.extend(["", ""])

        # anders dan gemeente
        agems = lookups["andersdangemeente"].get(iid, [])
        for i in range(max_anders_gemeente):
            row.append(agems[i] if i < len(agems) else "")

        # organisaties
        orgs = lookups["initiatieforganisatie"].get(iid, [])
        for i in range(max_organisaties):
            if i < len(orgs):
                o = lookups["organisaties"].get(orgs[i]["organisatie_id"], {})
                row.append(o.get("naam", f"org#{orgs[i]['organisatie_id']}"))
                row.append(orgs[i].get("rol", ""))
            else:
                row.extend(["", ""])

        # betrokken organisatietypen
        bets = lookups["betrokkenorganisatie"].get(iid, [])
        for i in range(max_betrokken):
            row.append(bets[i] if i < len(bets) else "")

        # api standaarden
        apis = lookups["initiatiefapistandaard"].get(iid, [])
        for i in range(max_apistandaarden):
            if i < len(apis):
                a = lookups["apistandaarden"].get(apis[i], {})
                row.append(a.get("naam", f"api#{apis[i]}"))
            else:
                row.append("")

        # andere api standaarden
        aapis = lookups["andereapistandaard"].get(iid, [])
        for i in range(max_andere_api):
            row.append(aapis[i] if i < len(aapis) else "")

        rows.append(row)

    write_sheet(ws, headers, rows, HEADER_FILL_ORANGE)

    # ── Contactpersoon ───────────────────────────────────────────────────

    ws = wb.create_sheet("Contactpersoon")
    rows = []
    for cp in lookups["contactpersonen"]:
        org = lookups["organisaties"].get(cp["organisatie_id"], {})
        pers = lookups["personen"].get(cp["persoon_id"], {})
        rows.append([
            cp["organisatie_id"],
            org.get("naam", ""),
            cp["persoon_id"],
            pers.get("naam", ""),
            pers.get("email", ""),
            cp.get("rol", ""),
        ])
    write_sheet(
        ws,
        ["organisatie_id", "organisatie.naam", "persoon_id", "persoon.naam", "persoon.email", "rol"],
        rows,
        HEADER_FILL_PURPLE,
    )

    return wb


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    print("Replay files laden…")
    lookups = build_lookups()
    print(f"  Gemeenten:       {len(lookups['gemeenten'])}")
    print(f"  Domeinen:        {len(lookups['domeinen'])}")
    print(f"  API standaarden: {len(lookups['apistandaarden'])}")
    print(f"  Organisaties:    {len(lookups['organisaties'])}")
    print(f"  Personen:        {len(lookups['personen'])}")
    print(f"  Initiatieven:    {len(lookups['initiatief_ids'])}")
    print(f"  Contactpersonen: {len(lookups['contactpersonen'])}")
    print()

    wb = build_excel(lookups)
    out = SCRIPT_DIR / "CG_Portfolio_Replay_Data.xlsx"
    wb.save(str(out))
    print(f"Excel opgeslagen: {out}")


if __name__ == "__main__":
    main()
